"""
Import/Export Service for Academic Management System

Handles bulk data import/export operations with Excel and CSV formats.
Supports users, students, courses, enrollments, and grades.
"""

from typing import List, Dict, Any, Optional, BinaryIO
from datetime import datetime
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
import bcrypt

from app.models import (
    User, Course, Semester, Enrollment, Grade, CourseSection as Schedule,
    Campus, Major
)
from app.models.user import UserRole
from app.services.username_generator import UsernameGenerator
from app.core.firebase import FirebaseService


class ImportExportService:
    """Service for handling data import and export operations"""
    
    def __init__(self):
        self.supported_formats = ['xlsx', 'csv']
        self.max_batch_size = 1000
    
    # ==================== VALIDATION ====================
    
    async def validate_import_data(
        self,
        data: pd.DataFrame,
        entity_type: str,
        db: AsyncSession
    ) -> Dict[str, Any]:
        """
        Validate import data before processing
        
        Returns:
            {
                "valid": bool,
                "errors": List[str],
                "warnings": List[str],
                "row_count": int
            }
        """
        errors = []
        warnings = []
        
        if data.empty:
            errors.append("Import file is empty")
            return {
                "valid": False,
                "errors": errors,
                "warnings": warnings,
                "row_count": 0
            }
        
        # Check batch size
        if len(data) > self.max_batch_size:
            errors.append(f"Import exceeds maximum batch size of {self.max_batch_size} rows")
        
        # Entity-specific validation
        if entity_type == "users":
            errors.extend(await self._validate_users(data, db))
        elif entity_type == "students":
            errors.extend(await self._validate_students(data, db))
        elif entity_type == "courses":
            errors.extend(await self._validate_courses(data, db))
        elif entity_type == "enrollments":
            errors.extend(await self._validate_enrollments(data, db))
        elif entity_type == "grades":
            errors.extend(await self._validate_grades(data, db))
        else:
            errors.append(f"Unsupported entity type: {entity_type}")
        
        # Check for duplicate rows
        duplicate_count = len(data) - len(data.drop_duplicates())
        if duplicate_count > 0:
            warnings.append(f"Found {duplicate_count} duplicate rows (will be skipped)")
        
        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": warnings,
            "row_count": len(data)
        }
    
    async def _validate_users(self, data: pd.DataFrame, db: AsyncSession) -> List[str]:
        """Validate user import data"""
        errors = []
        required_columns = ['full_name', 'role']
        
        # Check required columns
        missing_cols = [col for col in required_columns if col not in data.columns]
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return errors
        
        # Validate email format (if provided)
        if 'email' in data.columns:
            # Filter out empty/null emails
            emails_to_validate = data[data['email'].notna() & (data['email'] != '')]
            invalid_emails = emails_to_validate[~emails_to_validate['email'].str.contains('@', na=False)]
            if not invalid_emails.empty:
                rows = [idx + 2 for idx in invalid_emails.index.tolist()]  # +2 for header and 0-indexing
                errors.append(f"Invalid email format in rows: {rows}")
        
        # Validate role values
        valid_roles = [role.value for role in UserRole]
        invalid_roles = data[~data['role'].isin(valid_roles)]
        if not invalid_roles.empty:
            rows = [idx + 2 for idx in invalid_roles.index.tolist()]
            errors.append(f"Invalid role values in rows: {rows}. Valid roles: {', '.join(valid_roles)}")
        
        # Validate role-specific requirements
        for idx, row in data.iterrows():
            row_num = idx + 2
            role = row['role']
            
            # Students need campus_code, major_code, year_entered
            if role == 'student':
                if 'campus_code' not in data.columns or pd.isna(row.get('campus_code')):
                    errors.append(f"Row {row_num}: Students require campus_code")
                else:
                    # Validate campus exists
                    campus_code = str(row['campus_code']).strip().upper()
                    campus_result = await db.execute(select(Campus).where(Campus.code == campus_code))
                    if not campus_result.scalar_one_or_none():
                        errors.append(f"Row {row_num}: Campus code '{campus_code}' not found. Valid codes: H, D, C, S")
                
                if 'major_code' not in data.columns or pd.isna(row.get('major_code')):
                    errors.append(f"Row {row_num}: Students require major_code")
                else:
                    # Validate major exists
                    major_code = str(row['major_code']).strip().upper()
                    major_result = await db.execute(select(Major).where(Major.code == major_code))
                    if not major_result.scalar_one_or_none():
                        errors.append(f"Row {row_num}: Major code '{major_code}' not found")
                
                if 'year_entered' not in data.columns or pd.isna(row.get('year_entered')):
                    errors.append(f"Row {row_num}: Students require year_entered")
            
            # Teachers need campus_code
            elif role == 'teacher':
                if 'campus_code' not in data.columns or pd.isna(row.get('campus_code')):
                    errors.append(f"Row {row_num}: Teachers require campus_code")
                else:
                    # Validate campus exists
                    campus_code = str(row['campus_code']).strip().upper()
                    campus_result = await db.execute(select(Campus).where(Campus.code == campus_code))
                    if not campus_result.scalar_one_or_none():
                        errors.append(f"Row {row_num}: Campus code '{campus_code}' not found. Valid codes: H, D, C, S")
        
        return errors
    
    async def _validate_students(self, data: pd.DataFrame, db: AsyncSession) -> List[str]:
        """Validate student import data"""
        errors = []
        required_columns = ['user_id', 'student_id', 'date_of_birth']
        
        # Check required columns
        missing_cols = [col for col in required_columns if col not in data.columns]
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return errors
        
        # Check for duplicate student IDs
        duplicate_ids = data[data.duplicated(subset=['student_id'], keep=False)]
        if not duplicate_ids.empty:
            errors.append(f"Duplicate student IDs in rows: {duplicate_ids.index.tolist()}")
        
        # Validate user_id exists
        user_ids = data['user_id'].unique()
        result = await db.execute(select(User.id).where(User.id.in_(user_ids)))
        existing_user_ids = {row[0] for row in result.fetchall()}
        missing_user_ids = set(user_ids) - existing_user_ids
        
        if missing_user_ids:
            errors.append(f"User IDs not found in database: {list(missing_user_ids)[:5]}")
        
        return errors
    
    async def _validate_courses(self, data: pd.DataFrame, db: AsyncSession) -> List[str]:
        """Validate course import data"""
        errors = []
        required_columns = ['course_code', 'course_name', 'credits', 'department_id']
        
        # Check required columns
        missing_cols = [col for col in required_columns if col not in data.columns]
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return errors
        
        # Check for duplicate course codes
        duplicate_codes = data[data.duplicated(subset=['course_code'], keep=False)]
        if not duplicate_codes.empty:
            errors.append(f"Duplicate course codes in rows: {duplicate_codes.index.tolist()}")
        
        # Validate department_id exists
        dept_ids = data['department_id'].unique()
        result = await db.execute(select(Department.id).where(Department.id.in_(dept_ids)))
        existing_dept_ids = {row[0] for row in result.fetchall()}
        missing_dept_ids = set(dept_ids) - existing_dept_ids
        
        if missing_dept_ids:
            errors.append(f"Department IDs not found in database: {list(missing_dept_ids)}")
        
        # Validate credits is positive number
        if (data['credits'] <= 0).any():
            errors.append("Credits must be positive numbers")
        
        return errors
    
    async def _validate_enrollments(self, data: pd.DataFrame, db: AsyncSession) -> List[str]:
        """Validate enrollment import data"""
        errors = []
        required_columns = ['student_id', 'schedule_id']
        
        # Check required columns
        missing_cols = [col for col in required_columns if col not in data.columns]
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return errors
        
        # Validate student_id exists
        student_ids = data['student_id'].unique()
        result = await db.execute(select(Student.id).where(Student.id.in_(student_ids)))
        existing_student_ids = {row[0] for row in result.fetchall()}
        missing_student_ids = set(student_ids) - existing_student_ids
        
        if missing_student_ids:
            errors.append(f"Student IDs not found in database: {list(missing_student_ids)[:5]}")
        
        # Validate schedule_id exists
        schedule_ids = data['schedule_id'].unique()
        result = await db.execute(select(Schedule.id).where(Schedule.id.in_(schedule_ids)))
        existing_schedule_ids = {row[0] for row in result.fetchall()}
        missing_schedule_ids = set(schedule_ids) - existing_schedule_ids
        
        if missing_schedule_ids:
            errors.append(f"Schedule IDs not found in database: {list(missing_schedule_ids)[:5]}")
        
        return errors
    
    async def _validate_grades(self, data: pd.DataFrame, db: AsyncSession) -> List[str]:
        """Validate grade import data"""
        errors = []
        required_columns = ['enrollment_id', 'grade_value']
        
        # Check required columns
        missing_cols = [col for col in required_columns if col not in data.columns]
        if missing_cols:
            errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return errors
        
        # Validate enrollment_id exists
        enrollment_ids = data['enrollment_id'].unique()
        result = await db.execute(select(Enrollment.id).where(Enrollment.id.in_(enrollment_ids)))
        existing_enrollment_ids = {row[0] for row in result.fetchall()}
        missing_enrollment_ids = set(enrollment_ids) - existing_enrollment_ids
        
        if missing_enrollment_ids:
            errors.append(f"Enrollment IDs not found in database: {list(missing_enrollment_ids)[:5]}")
        
        # Validate grade values (0-100 or letter grades)
        if 'grade_value' in data.columns:
            invalid_grades = data[
                ~data['grade_value'].astype(str).str.match(r'^(\d+(\.\d+)?|[A-F][+-]?|Pass|Fail)$', na=False)
            ]
            if not invalid_grades.empty:
                errors.append(f"Invalid grade values in rows: {invalid_grades.index.tolist()}")
        
        return errors
    
    # ==================== IMPORT OPERATIONS ====================
    
    async def import_users(
        self,
        data: pd.DataFrame,
        db: AsyncSession,
        skip_existing: bool = True
    ) -> Dict[str, Any]:
        """Import users from DataFrame with automatic username generation"""
        imported = 0
        skipped = 0
        errors = []
        
        for idx, row in data.iterrows():
            row_num = idx + 2  # +2 for header and 0-indexing
            try:
                full_name = row['full_name']
                role = row['role']
                
                # Generate username based on role
                if role == 'student':
                    campus_code = str(row['campus_code']).strip().upper()
                    major_code = str(row['major_code']).strip().upper()
                    year_entered = int(row['year_entered'])
                    
                    # Fetch campus and major to get IDs
                    campus_result = await db.execute(select(Campus).where(Campus.code == campus_code))
                    campus = campus_result.scalar_one_or_none()
                    if not campus:
                        errors.append(f"Row {row_num}: Campus code '{campus_code}' not found")
                        continue
                    
                    major_result = await db.execute(select(Major).where(Major.code == major_code))
                    major = major_result.scalar_one_or_none()
                    if not major:
                        errors.append(f"Row {row_num}: Major code '{major_code}' not found")
                        continue
                    
                    username = await UsernameGenerator.generate_student_username(
                        db, full_name, major.code, campus.code, year_entered
                    )
                elif role == 'teacher':
                    campus_code = str(row['campus_code']).strip().upper()
                    
                    # Fetch campus to get ID
                    campus_result = await db.execute(select(Campus).where(Campus.code == campus_code))
                    campus = campus_result.scalar_one_or_none()
                    if not campus:
                        errors.append(f"Row {row_num}: Campus code '{campus_code}' not found")
                        continue
                    
                    username = await UsernameGenerator.generate_teacher_username(
                        db, full_name, campus.code
                    )
                else:
                    # Admin, staff, etc. - still need campus for staff usernames
                    campus_code = str(row.get('campus_code', 'H')).strip().upper()  # Default to Hanoi if not provided
                    
                    # Fetch campus to get ID
                    campus_result = await db.execute(select(Campus).where(Campus.code == campus_code))
                    campus = campus_result.scalar_one_or_none()
                    if not campus:
                        errors.append(f"Row {row_num}: Campus code '{campus_code}' not found")
                        continue
                    
                    username = await UsernameGenerator.generate_staff_username(
                        db, full_name, campus.code, role
                    )
                
                # Check if username already exists (shouldn't happen with generator, but safety check)
                result = await db.execute(select(User).where(User.username == username))
                if result.scalar_one_or_none():
                    if skip_existing:
                        skipped += 1
                        continue
                    else:
                        errors.append(f"Row {row_num}: Username {username} already exists")
                        continue
                
                # Generate Greenwich email for Firebase
                firebase_email = UsernameGenerator.generate_email(username, role)
                
                # Create Firebase user
                try:
                    firebase_user = FirebaseService.create_user(
                        email=firebase_email,
                        password=username,  # Default password = username
                        display_name=full_name
                    )
                    firebase_uid = firebase_user.uid
                    
                    # Set custom claims for role-based access
                    custom_claims = {
                        "role": role,
                        "username": username
                    }
                    FirebaseService.set_custom_user_claims(firebase_uid, custom_claims)
                    
                except Exception as firebase_error:
                    errors.append(f"Row {row_num}: Firebase error - {str(firebase_error)}")
                    continue
                
                # Create database user
                # Convert phone_number to string if it exists and is not empty
                phone_number = row.get('phone_number')
                if pd.notna(phone_number) and phone_number != '':
                    phone_number = str(phone_number).strip()
                else:
                    phone_number = None
                
                user_data = {
                    'username': username,
                    'email': row.get('email') if pd.notna(row.get('email')) and row.get('email') != '' else None,
                    'full_name': full_name,
                    'role': role,
                    'phone_number': phone_number,
                    'campus_id': campus.id if campus else None,
                    'firebase_uid': firebase_uid,
                    'status': 'active'
                }
                
                # Add role-specific fields
                if role == 'student' and major:
                    user_data['major_id'] = major.id
                    user_data['year_entered'] = int(row['year_entered'])
                
                user = User(**user_data)
                db.add(user)
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        try:
            await db.commit()
        except Exception as e:
            await db.rollback()
            errors.append(f"Database commit error: {str(e)}")
        
        return {
            "imported": imported,
            "skipped": skipped,
            "errors": errors,
            "total_rows": len(data)
        }
    
    async def import_students(
        self,
        data: pd.DataFrame,
        db: AsyncSession,
        skip_existing: bool = True
    ) -> Dict[str, Any]:
        """Import students from DataFrame"""
        imported = 0
        skipped = 0
        errors = []
        
        # Remove duplicates
        data = data.drop_duplicates(subset=['student_id'])
        
        for idx, row in data.iterrows():
            try:
                # Check if student exists
                if skip_existing:
                    result = await db.execute(select(Student).where(Student.student_id == row['student_id']))
                    if result.scalar_one_or_none():
                        skipped += 1
                        continue
                
                # Parse date
                date_of_birth = pd.to_datetime(row['date_of_birth']).date()
                
                # Create student
                student = Student(
                    user_id=row['user_id'],
                    student_id=row['student_id'],
                    date_of_birth=date_of_birth,
                    gender=row.get('gender'),
                    admission_date=pd.to_datetime(row['admission_date']).date() if 'admission_date' in row and pd.notna(row['admission_date']) else None,
                    graduation_date=pd.to_datetime(row['graduation_date']).date() if 'graduation_date' in row and pd.notna(row['graduation_date']) else None,
                    status=row.get('status', 'active')
                )
                db.add(student)
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        await db.commit()
        
        return {
            "imported": imported,
            "skipped": skipped,
            "errors": errors,
            "total_rows": len(data)
        }
    
    async def import_courses(
        self,
        data: pd.DataFrame,
        db: AsyncSession,
        skip_existing: bool = True
    ) -> Dict[str, Any]:
        """Import courses from DataFrame"""
        imported = 0
        skipped = 0
        errors = []
        
        # Remove duplicates
        data = data.drop_duplicates(subset=['course_code'])
        
        for idx, row in data.iterrows():
            try:
                # Check if course exists
                if skip_existing:
                    result = await db.execute(select(Course).where(Course.course_code == row['course_code']))
                    if result.scalar_one_or_none():
                        skipped += 1
                        continue
                
                # Create course
                course = Course(
                    course_code=row['course_code'],
                    course_name=row['course_name'],
                    description=row.get('description'),
                    credits=int(row['credits']),
                    department_id=int(row['department_id']),
                    is_active=row.get('is_active', True)
                )
                db.add(course)
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        await db.commit()
        
        return {
            "imported": imported,
            "skipped": skipped,
            "errors": errors,
            "total_rows": len(data)
        }
    
    async def import_enrollments(
        self,
        data: pd.DataFrame,
        db: AsyncSession,
        skip_existing: bool = True
    ) -> Dict[str, Any]:
        """Import enrollments from DataFrame"""
        imported = 0
        skipped = 0
        errors = []
        
        # Remove duplicates
        data = data.drop_duplicates(subset=['student_id', 'schedule_id'])
        
        for idx, row in data.iterrows():
            try:
                # Check if enrollment exists
                if skip_existing:
                    result = await db.execute(
                        select(Enrollment).where(
                            and_(
                                Enrollment.student_id == row['student_id'],
                                Enrollment.schedule_id == row['schedule_id']
                            )
                        )
                    )
                    if result.scalar_one_or_none():
                        skipped += 1
                        continue
                
                # Create enrollment
                enrollment = Enrollment(
                    student_id=int(row['student_id']),
                    schedule_id=int(row['schedule_id']),
                    enrollment_date=pd.to_datetime(row['enrollment_date']).date() if 'enrollment_date' in row and pd.notna(row['enrollment_date']) else datetime.now().date(),
                    status=row.get('status', 'enrolled')
                )
                db.add(enrollment)
                imported += 1
                
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        await db.commit()
        
        return {
            "imported": imported,
            "skipped": skipped,
            "errors": errors,
            "total_rows": len(data)
        }
    
    # ==================== EXPORT OPERATIONS ====================
    
    async def export_users(
        self,
        db: AsyncSession,
        format: str = 'xlsx',
        role_filter: Optional[str] = None
    ) -> bytes:
        """Export users to Excel or CSV"""
        query = select(User)
        
        if role_filter:
            query = query.where(User.role == role_filter)
        
        result = await db.execute(query)
        users = result.scalars().all()
        
        # Convert to DataFrame
        data = []
        for user in users:
            data.append({
                'id': user.id,
                'email': user.email,
                'full_name': user.full_name,
                'role': user.role,
                'is_active': user.is_active,
                'phone_number': user.phone_number,
                'address': user.address,
                'created_at': user.created_at
            })
        
        df = pd.DataFrame(data)
        
        if format == 'csv':
            return df.to_csv(index=False).encode('utf-8')
        else:
            return self._dataframe_to_excel(df, 'Users')
    
    async def export_students(
        self,
        db: AsyncSession,
        format: str = 'xlsx',
        status_filter: Optional[str] = None
    ) -> bytes:
        """Export students with user details to Excel or CSV"""
        query = select(Student, User).join(User, Student.user_id == User.id)
        
        if status_filter:
            query = query.where(Student.status == status_filter)
        
        result = await db.execute(query)
        students = result.all()
        
        # Convert to DataFrame
        data = []
        for student, user in students:
            data.append({
                'id': student.id,
                'student_id': student.student_id,
                'full_name': user.full_name,
                'email': user.email,
                'date_of_birth': student.date_of_birth,
                'gender': student.gender,
                'admission_date': student.admission_date,
                'graduation_date': student.graduation_date,
                'status': student.status,
                'phone_number': user.phone_number,
                'address': user.address
            })
        
        df = pd.DataFrame(data)
        
        if format == 'csv':
            return df.to_csv(index=False).encode('utf-8')
        else:
            return self._dataframe_to_excel(df, 'Students')
    
    async def export_courses(
        self,
        db: AsyncSession,
        format: str = 'xlsx',
        department_id: Optional[int] = None
    ) -> bytes:
        """Export courses with department details to Excel or CSV"""
        query = select(Course, Department).join(Department, Course.department_id == Department.id)
        
        if department_id:
            query = query.where(Course.department_id == department_id)
        
        result = await db.execute(query)
        courses = result.all()
        
        # Convert to DataFrame
        data = []
        for course, department in courses:
            data.append({
                'id': course.id,
                'course_code': course.course_code,
                'course_name': course.course_name,
                'description': course.description,
                'credits': course.credits,
                'department_name': department.name,
                'department_id': department.id,
                'is_active': course.is_active
            })
        
        df = pd.DataFrame(data)
        
        if format == 'csv':
            return df.to_csv(index=False).encode('utf-8')
        else:
            return self._dataframe_to_excel(df, 'Courses')
    
    async def export_enrollments(
        self,
        db: AsyncSession,
        format: str = 'xlsx',
        semester_id: Optional[int] = None
    ) -> bytes:
        """Export enrollments with student and course details to Excel or CSV"""
        query = (
            select(Enrollment, Student, User, Schedule, Course)
            .join(Student, Enrollment.student_id == Student.id)
            .join(User, Student.user_id == User.id)
            .join(Schedule, Enrollment.schedule_id == Schedule.id)
            .join(Course, Schedule.course_id == Course.id)
        )
        
        if semester_id:
            query = query.where(Schedule.semester_id == semester_id)
        
        result = await db.execute(query)
        enrollments = result.all()
        
        # Convert to DataFrame
        data = []
        for enrollment, student, user, schedule, course in enrollments:
            data.append({
                'enrollment_id': enrollment.id,
                'student_id': student.student_id,
                'student_name': user.full_name,
                'course_code': course.course_code,
                'course_name': course.course_name,
                'schedule_id': schedule.id,
                'enrollment_date': enrollment.enrollment_date,
                'status': enrollment.status
            })
        
        df = pd.DataFrame(data)
        
        if format == 'csv':
            return df.to_csv(index=False).encode('utf-8')
        else:
            return self._dataframe_to_excel(df, 'Enrollments')
    
    async def export_grades(
        self,
        db: AsyncSession,
        format: str = 'xlsx',
        semester_id: Optional[int] = None,
        student_id: Optional[int] = None
    ) -> bytes:
        """Export grades with student and course details to Excel or CSV"""
        query = (
            select(Grade, Enrollment, Student, User, Schedule, Course)
            .join(Enrollment, Grade.enrollment_id == Enrollment.id)
            .join(Student, Enrollment.student_id == Student.id)
            .join(User, Student.user_id == User.id)
            .join(Schedule, Enrollment.schedule_id == Schedule.id)
            .join(Course, Schedule.course_id == Course.id)
        )
        
        if semester_id:
            query = query.where(Schedule.semester_id == semester_id)
        
        if student_id:
            query = query.where(Student.id == student_id)
        
        result = await db.execute(query)
        grades = result.all()
        
        # Convert to DataFrame
        data = []
        for grade, enrollment, student, user, schedule, course in grades:
            data.append({
                'grade_id': grade.id,
                'student_id': student.student_id,
                'student_name': user.full_name,
                'course_code': course.course_code,
                'course_name': course.course_name,
                'grade_type': grade.grade_type,
                'grade_value': grade.grade_value,
                'weight': grade.weight,
                'graded_by': grade.graded_by_id,
                'graded_at': grade.graded_at,
                'is_published': grade.is_published
            })
        
        df = pd.DataFrame(data)
        
        if format == 'csv':
            return df.to_csv(index=False).encode('utf-8')
        else:
            return self._dataframe_to_excel(df, 'Grades')
    
    # ==================== TEMPLATE GENERATION ====================
    
    def generate_import_template(self, entity_type: str, format: str = 'xlsx') -> bytes:
        """Generate import template with sample data and instructions"""
        
        templates = {
            'users': pd.DataFrame([
                {
                    'email': 'john.doe@example.com',
                    'full_name': 'John Doe',
                    'role': 'student',
                    'password': 'changeme123',
                    'is_active': True,
                    'phone_number': '+1234567890',
                    'address': '123 Main St'
                }
            ]),
            'students': pd.DataFrame([
                {
                    'user_id': 1,
                    'student_id': 'S2024001',
                    'date_of_birth': '2000-01-15',
                    'gender': 'male',
                    'admission_date': '2024-09-01',
                    'status': 'active'
                }
            ]),
            'courses': pd.DataFrame([
                {
                    'course_code': 'CS101',
                    'course_name': 'Introduction to Computer Science',
                    'description': 'Fundamentals of programming',
                    'credits': 3,
                    'department_id': 1,
                    'is_active': True
                }
            ]),
            'enrollments': pd.DataFrame([
                {
                    'student_id': 1,
                    'schedule_id': 1,
                    'enrollment_date': '2024-09-01',
                    'status': 'enrolled'
                }
            ]),
            'grades': pd.DataFrame([
                {
                    'enrollment_id': 1,
                    'grade_type': 'midterm',
                    'grade_value': 85.5,
                    'weight': 30.0
                }
            ])
        }
        
        if entity_type not in templates:
            raise ValueError(f"Unsupported entity type: {entity_type}")
        
        df = templates[entity_type]
        
        if format == 'csv':
            return df.to_csv(index=False).encode('utf-8')
        else:
            return self._dataframe_to_excel(df, f'{entity_type.capitalize()} Import Template', is_template=True)
    
    # ==================== HELPER METHODS ====================
    
    def _dataframe_to_excel(
        self,
        df: pd.DataFrame,
        sheet_name: str,
        is_template: bool = False
    ) -> bytes:
        """Convert DataFrame to Excel file with formatting"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Style header row
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add instructions sheet for templates
            if is_template:
                instructions_sheet = workbook.create_sheet('Instructions', 0)
                instructions = [
                    ['Import Template Instructions'],
                    [''],
                    ['1. Fill in the data in the template sheet'],
                    ['2. Do not modify the column headers'],
                    ['3. Remove the sample row before importing'],
                    ['4. Make sure all required fields are filled'],
                    ['5. Save the file and upload for import'],
                    [''],
                    ['Required Fields:'],
                    ['- Fields cannot be empty'],
                    ['- Check the column headers for required fields'],
                    [''],
                    ['Data Format:'],
                    ['- Dates: YYYY-MM-DD (e.g., 2024-09-01)'],
                    ['- Booleans: TRUE or FALSE'],
                    ['- Numbers: No commas or special characters']
                ]
                
                for row_idx, row_data in enumerate(instructions, 1):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = instructions_sheet.cell(row=row_idx, column=col_idx, value=value)
                        if row_idx == 1:
                            cell.font = Font(bold=True, size=14)
                        elif value.startswith(('1.', '2.', '3.', '4.', '5.')):
                            cell.font = Font(bold=True)
        
        output.seek(0)
        return output.read()
    
    def parse_upload_file(
        self,
        file: BinaryIO,
        filename: str
    ) -> pd.DataFrame:
        """Parse uploaded file (Excel or CSV) into DataFrame"""
        file_extension = filename.split('.')[-1].lower()
        
        if file_extension not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {file_extension}. Supported: {', '.join(self.supported_formats)}")
        
        # Read file content
        content = file.read()
        
        if file_extension == 'csv':
            df = pd.read_csv(io.BytesIO(content))
        else:  # xlsx
            df = pd.read_excel(io.BytesIO(content), engine='openpyxl')
        
        # Clean column names (strip whitespace)
        df.columns = df.columns.str.strip()
        
        # Remove empty rows
        df = df.dropna(how='all')
        
        return df


# Create service instance
import_export_service = ImportExportService()
